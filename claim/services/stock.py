from openpyxl import load_workbook

class FileStockHandler():
    def __init__(self, file):    
        self.wb = load_workbook(file, read_only=True)
        self.ws = self.wb.active

    def create_iter(self):
        rows = self.ws.iter_rows(values_only=True)
        headers = next(rows)    
        header_idx = {h: i for i, h in enumerate(headers)}
        return rows, header_idx

    def oldest_symbols(self):
        rows, header_idx = self.create_iter()
        by_symbol = {}
        for row in rows:
            symbol = row[header_idx['Symbol']]
            trade_date = row[header_idx['Trade Date']]

            if (
                symbol is None 
                or str(symbol).strip()=='' 
                or row[header_idx['Quantity']] is None 
                or row[header_idx['Amount']] is None 
                or trade_date is None
            ):
                continue

            if symbol not in by_symbol:
                by_symbol[symbol] = row
                continue

            current = by_symbol[symbol]
            if trade_date < current[header_idx['Trade Date']]:
                by_symbol[symbol] = row
        return by_symbol

    def rows_by_symbols(self, symbols):
        rows, header_idx = self.create_iter()
        symbols = {str(s).strip() for s in symbols} 
        for row in rows:
            symbol = row[header_idx["Symbol"]]
            if symbol is None:
                continue
            symbol = str(symbol).strip()
            if symbol not in symbols:
                continue
            yield row



